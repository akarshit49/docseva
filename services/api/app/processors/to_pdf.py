#genai: Export doc/docx/xls/xlsx to PDF — pure Python via fpdf2, no system tools.
from __future__ import annotations

from pathlib import Path


def convert_to_pdf(input_path: Path, output_path: Path) -> Path:
    """
    Convert input_path to PDF and write the result to output_path.
    Uses only pure-Python open-source libraries (fpdf2, openpyxl, xlrd).
    No LibreOffice or other external tools are required.
    """
    suffix = input_path.suffix.lower()
    if suffix in {".doc", ".docx"}:
        _docx_to_pdf(input_path, output_path)
    elif suffix in {".xlsx", ".xls"}:
        _excel_to_pdf(input_path, output_path)
    else:
        raise ValueError(f"Cannot convert {suffix} to PDF. Supported: .doc .docx .xlsx .xls")

    return output_path


# ---------------------------------------------------------------------------
# DOCX / DOC → PDF
# ---------------------------------------------------------------------------

def _docx_to_pdf(input_path: Path, output_path: Path) -> None:
    """
    Render doc/docx as a PDF preserving table structure.
    Reads the file with python-docx (for .docx) or olefile (for .doc),
    then generates a professional PDF with fpdf2.
    """
    from fpdf import FPDF
    from app.processors.extractors import extract_text

    suffix = input_path.suffix.lower()

    if suffix == ".docx":
        # Rich path: use python-docx to read tables and paragraphs separately.
        from docx import Document as DocxDocument
        doc = DocxDocument(str(input_path))
        pdf = _make_pdf()
        pdf.add_page()

        _pdf_title(pdf, input_path.stem)

        for block in doc.element.body:
            tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag
            if tag == "p":
                from docx.oxml.ns import qn as _qn
                texts = [n.text for n in block.iter() if n.tag == _qn("w:t") and n.text]
                line = "".join(texts).strip()
                if line:
                    _pdf_para(pdf, line)
            elif tag == "tbl":
                # Extract table rows
                from docx.oxml.ns import qn as _qn
                rows_data: list[list[str]] = []
                for tr in block.findall(f".//{_qn('w:tr')}"):
                    cells = []
                    for tc in tr.findall(f".//{_qn('w:tc')}"):
                        texts = [n.text for n in tc.iter() if n.tag == _qn("w:t") and n.text]
                        cells.append("".join(texts).strip())
                    if cells:
                        rows_data.append(cells)
                if rows_data:
                    _pdf_table(pdf, rows_data)
    else:
        # .doc — plain text extraction with olefile
        text = extract_text(str(input_path))
        pdf = _make_pdf()
        pdf.add_page()
        _pdf_title(pdf, input_path.stem)
        for line in text.split("\n"):
            _pdf_para(pdf, line)

    pdf.output(str(output_path))


# ---------------------------------------------------------------------------
# Excel → PDF
# ---------------------------------------------------------------------------

def _excel_to_pdf(input_path: Path, output_path: Path) -> None:
    suffix = input_path.suffix.lower()
    if suffix == ".xlsx":
        sheets = _load_xlsx(input_path)
    else:
        sheets = _load_xls(input_path)

    from fpdf import FPDF
    pdf = _make_pdf()

    for idx, (sheet_name, rows) in enumerate(sheets):
        if idx > 0:
            pdf.add_page()
        else:
            pdf.add_page()

        pdf.set_font("Helvetica", "B", 13)
        _safe_cell(pdf, 0, 8, sheet_name, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

        if rows:
            _pdf_table(pdf, [[_cell_str(v) for v in row] for row in rows])

    pdf.output(str(output_path))


def _load_xlsx(path: Path) -> list[tuple[str, list[list]]]:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows():
            row_data = [cell.value for cell in row]
            if any(v is not None for v in row_data):
                rows.append(row_data)
        sheets.append((name, rows))
    return sheets


def _load_xls(path: Path) -> list[tuple[str, list[list]]]:
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sheets = []
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        rows = []
        for r in range(ws.nrows):
            row_data = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if any(v != "" and v is not None for v in row_data):
                rows.append(row_data)
        sheets.append((name, rows))
    return sheets


# ---------------------------------------------------------------------------
# fpdf2 helpers
# ---------------------------------------------------------------------------

def _make_pdf():
    from fpdf import FPDF
    pdf = FPDF()
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    return pdf


def _pdf_title(pdf, title: str) -> None:
    pdf.set_font("Helvetica", "B", 14)
    _safe_cell(pdf, 0, 9, title, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)


def _pdf_para(pdf, text: str) -> None:
    if not text.strip():
        pdf.ln(3)
        return
    pdf.set_font("Helvetica", "", 10)
    # Reset X to left margin in case a previous cell left us mid-line.
    try:
        pdf.set_x(pdf.l_margin)
        _safe_multi_cell(pdf, 0, 6, text)
    except Exception:
        # Fall back to truncated single-line rendering if multi_cell can't fit
        # an oversized token. Better to print less than crash the whole conversion.
        pdf.set_x(pdf.l_margin)
        truncated = text[:200] + ("…" if len(text) > 200 else "")
        try:
            _safe_cell(pdf, 0, 6, truncated, new_x="LMARGIN", new_y="NEXT")
        except Exception:
            pass


#genai: Content-aware column widths + text wrapping to prevent overlap
def _pdf_table(pdf, rows: list[list[str]]) -> None:
    if not rows:
        return
    #genai: Defensive — ensure table starts at the left margin regardless of prior cursor state
    pdf.set_x(pdf.l_margin)
    usable = pdf.w - pdf.l_margin - pdf.r_margin
    max_cols = max(len(r) for r in rows)

    pdf.set_font("Helvetica", "", 9)
    col_widths = _calc_col_widths(pdf, rows, max_cols, usable)

    for i, row in enumerate(rows):
        is_header = i == 0
        pdf.set_font("Helvetica", "B" if is_header else "", 9)
        row_h = _render_table_row(pdf, row, col_widths, max_cols, is_header)
    pdf.ln(3)


def _calc_col_widths(pdf, rows: list[list[str]], max_cols: int, usable: float) -> list[float]:
    """Compute column widths proportional to content, with sensible min/max."""
    padding = 3
    max_w = [0.0] * max_cols
    for row in rows:
        for j in range(max_cols):
            val = row[j] if j < len(row) else ""
            tw = pdf.get_string_width(_safe(str(val))) + padding
            max_w[j] = max(max_w[j], tw)

    total = sum(max_w) or 1
    col_widths = [w / total * usable for w in max_w]

    min_col = 10
    max_col_pct = 0.55
    for j in range(max_cols):
        col_widths[j] = max(col_widths[j], min_col)
        col_widths[j] = min(col_widths[j], usable * max_col_pct)

    # Re-normalise so columns sum to exactly usable width
    scale = usable / (sum(col_widths) or 1)
    return [w * scale for w in col_widths]


def _render_table_row(
    pdf, row: list[str], col_widths: list[float], max_cols: int, is_header: bool,
) -> float:
    """Render a single table row, wrapping text if it exceeds column width."""
    base_h = 6
    x_start = pdf.get_x()
    y_start = pdf.get_y()

    # First pass: compute the height needed for the tallest cell
    cell_lines: list[list[str]] = []
    max_lines = 1
    for j in range(max_cols):
        val = _safe(str(row[j]) if j < len(row) else "")
        lines = _wrap_text(pdf, val, col_widths[j] - 2)
        cell_lines.append(lines)
        max_lines = max(max_lines, len(lines))

    row_h = base_h * max_lines

    # Page break check
    if pdf.get_y() + row_h > pdf.h - pdf.b_margin:
        pdf.add_page()
        y_start = pdf.get_y()

    # Second pass: draw each cell
    for j in range(max_cols):
        x = x_start + sum(col_widths[:j])
        pdf.rect(x, y_start, col_widths[j], row_h)
        lines = cell_lines[j]
        for k, line in enumerate(lines):
            pdf.set_xy(x + 1, y_start + k * base_h)
            pdf.cell(col_widths[j] - 2, base_h, line)

    pdf.set_xy(x_start, y_start + row_h)
    return row_h


def _wrap_text(pdf, text: str, max_width: float) -> list[str]:
    """Break text into lines that fit within max_width."""
    if not text:
        return [""]
    if pdf.get_string_width(text) <= max_width:
        return [text]

    lines: list[str] = []
    words = text.split()
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip() if current else word
        if pdf.get_string_width(candidate) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            # If a single word exceeds the width, force-break it
            if pdf.get_string_width(word) > max_width:
                while word:
                    for end in range(len(word), 0, -1):
                        if pdf.get_string_width(word[:end]) <= max_width:
                            lines.append(word[:end])
                            word = word[end:]
                            break
                    else:
                        lines.append(word[:1])
                        word = word[1:]
                current = ""
            else:
                current = word
    if current:
        lines.append(current)
    return lines or [""]


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val)


def _safe(text: str) -> str:
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _safe_cell(pdf, w, h, text, **kwargs):
    pdf.cell(w, h, _safe(str(text)), **kwargs)


def _safe_multi_cell(pdf, w, h, text):
    #genai: fpdf2 default leaves cursor at right edge; force return to left margin
    from fpdf.enums import XPos, YPos
    pdf.multi_cell(w, h, _safe(str(text)), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
