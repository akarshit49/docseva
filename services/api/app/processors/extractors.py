#genai: File text extraction for doc/docx/pdf — pure Python, no external tools.
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import olefile
from pypdf import PdfReader


def extract_text(file_path: str) -> str:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".docx":
        return _extract_docx_text(path)
    if suffix == ".pdf":
        return _extract_pdf_text(path)
    if suffix == ".doc":
        return _extract_doc_text(path)
    if suffix == ".xlsx":
        return _extract_xlsx_text(path)
    if suffix == ".xls":
        return _extract_xls_text(path)

    raise ValueError(
        f"Unsupported file type: {suffix}. "
        "Supported: .doc, .docx, .pdf, .xls, .xlsx"
    )


# ── Excel ────────────────────────────────────────────────────────────────────

def _extract_xlsx_text(path: Path) -> str:
    """Extract text from a .xlsx workbook. Concatenates sheets and rows."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"### Sheet: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_xls_text(path: Path) -> str:
    """Extract text from a legacy .xls workbook (Excel 97–2003) via xlrd."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    lines: list[str] = []
    for sheet_name in wb.sheet_names():
        sheet = wb.sheet_by_name(sheet_name)
        lines.append(f"### Sheet: {sheet_name}")
        for r in range(sheet.nrows):
            cells = []
            for c in range(sheet.ncols):
                v = sheet.cell_value(r, c)
                if isinstance(v, float) and v == int(v):
                    cells.append(str(int(v)))
                else:
                    cells.append("" if v in ("", None) else str(v))
            if any(cell.strip() for cell in cells):
                lines.append("\t".join(cells))
    return "\n".join(lines)


def _extract_doc_text(path: Path) -> str:
    """
    Extract text from a legacy .doc (Word Binary Format) file via pure-Python
    OLE stream parsing — no external tools (LibreOffice / antiword) required.
    """
    return _extract_doc_ole(path)


def _extract_doc_ole(path: Path) -> str:
    """
    Read text from .doc OLE streams without any external tool.
    Extracts printable ASCII from WordDocument stream and Unicode
    from the Table stream — sufficient for structured quotation tables.
    """
    if not olefile.isOleFile(str(path)):
        raise ValueError(f"File does not appear to be a valid .doc file: {path.name}")

    ole = olefile.OleFileIO(str(path))
    try:
        chunks: list[str] = []

        if ole.exists("WordDocument"):
            data = ole.openstream("WordDocument").read()
            ascii_chunks = re.findall(rb"[ -~\t\r\n]{4,}", data)
            chunks += [c.decode("ascii", errors="ignore").strip() for c in ascii_chunks]

        for stream_name in ("1Table", "0Table"):
            if ole.exists(stream_name):
                data = ole.openstream(stream_name).read()
                unicode_chunks = re.findall(rb"(?:[\x20-\x7e]\x00){4,}", data)
                chunks += [c.decode("utf-16-le", errors="ignore").strip() for c in unicode_chunks]
                break

        lines = [ln for ln in chunks if ln and not ln.startswith("bjbj")]
        return "\n".join(lines)
    finally:
        ole.close()


def _extract_docx_text(path: Path) -> str:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    lines: list[str] = []
    for paragraph in root.findall(".//w:p", ns):
        text_nodes = [node.text for node in paragraph.findall(".//w:t", ns) if node.text]
        line = "".join(text_nodes).strip()
        if line:
            lines.append(line)
    return "\n".join(lines)


def _extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(pages).strip()
